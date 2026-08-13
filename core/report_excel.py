"""xlsx review report builder — same data as the HTML report, color-coded rows, Audio hyperlink column."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

FILL_NOTE = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
FILL_NARRATOR = PatternFill(start_color="DBE9FF", end_color="DBE9FF", fill_type="solid")
FILL_HEADER = PatternFill(start_color="333333", end_color="333333", fill_type="solid")

HEADERS = ["#", "Chapter", "Speaker", "English", "Translated", "Reviewer Comment", "Flag", "Audio"]


def _row_fill(row: dict) -> PatternFill | None:
    if row["speaker"].lower() == "narrator":
        return FILL_NARRATOR
    if row.get("flag") == "note":
        return FILL_NOTE
    return None


def build_excel_report(parsed: dict, reviews: list[dict], audio_paths: dict[int, str],
                        target_language: str, output_path: str) -> None:
    rows = parsed["rows"]
    for row, review in zip(rows, reviews):
        row["comment"] = review["comment"]
        row["flag"] = review["flag"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Translation Review"

    ws.append([f"{parsed['title']} — Translation Review ({target_language.title()})"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    ws["A1"].font = Font(bold=True, size=14)

    ws.append([f"Theme: {parsed['theme']}"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))
    ws.append([])

    header_row_idx = 4
    ws.append(HEADERS)
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row in rows:
        ws.append([
            row["sr_no"], row["chapter"], row["speaker"],
            row["english"], row["translated"], row["comment"], row["flag"], "",
        ])
        excel_row_idx = ws.max_row
        fill = _row_fill(row)
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=excel_row_idx, column=col_idx)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill

        audio_file = audio_paths.get(row["sr_no"])
        audio_cell = ws.cell(row=excel_row_idx, column=len(HEADERS))
        if audio_file:
            audio_cell.value = "Play audio"
            audio_cell.hyperlink = f"audio/{audio_file}"
            audio_cell.font = Font(color="0563C1", underline="single")
        else:
            audio_cell.value = "no audio"

    widths = {1: 5, 2: 9, 3: 12, 4: 40, 5: 40, 6: 45, 7: 8, 8: 14}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = f"A{header_row_idx + 1}"
    wb.save(output_path)
